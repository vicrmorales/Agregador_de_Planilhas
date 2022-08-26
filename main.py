from pickle import FALSE, TRUE
import openpyxl as xl
from openpyxl import Workbook 
from os import listdir
from os.path import isfile, join
import pathlib

local_path = pathlib.Path(__file__).parent.resolve()
print(local_path)

dir_planilhas_base = str(local_path) + "/Planilhas Base"
planilhas_base = [f for f in listdir(dir_planilhas_base) if isfile(join(dir_planilhas_base, f))]
dir_arquivo_final = str(local_path) + '/Planilha_Agregada.xlsx'
lin_arquivo_final = 1
arquivo_final = Workbook()
arquivo_final_s = arquivo_final.active
bPrimeiraPlanilha = TRUE

for file in planilhas_base:
    filename = dir_planilhas_base + "/" + file
    wb1 = xl.load_workbook (filename)
    ws1 = wb1.worksheets [0]

    mr = ws1.max_row

    mc = ws1.max_column

    if bPrimeiraPlanilha == TRUE: # Copia do cabeçalho da primeira planilha base
        for x in range ( 1 , mc + 1 ):  
            c = ws1.cell (row = 2, column = x)  # O numero do row define a linha na qual o cabeçalho está.
            arquivo_final_s.cell (row = 1, column = x) .value = c.value
            #print(arquivo_final_s.cell (row = 1, column = x) .value)
        bPrimeiraPlanilha = FALSE
        lin_arquivo_final = 2

    for i in range ( 3 , mr + 1 ): # O primeiro campo define qual a linha na qual os dados que desejam ser agregados começam.
        c_test = ws1.cell (row = i, column = 1)
        if c_test.value is not None:
            for j in range ( 1 , mc + 1 ):  
                c = ws1.cell (row = i, column = j)  
                arquivo_final_s.cell (row = lin_arquivo_final, column = j) .value = c.value
            lin_arquivo_final = lin_arquivo_final + 1        

print(lin_arquivo_final)

arquivo_final.save (str(dir_arquivo_final)) 